from openpyxl import load_workbook
from openpyxl.workbook import Workbook
from typing import Dict, List, Any
import os
from datetime import datetime

class ExcelHandler:
    def __init__(self, template_path: str):
        self.template_path = template_path
        
    def populate_template(self, data: List[Dict[str, Any]], 
                         column_mappings: Dict[str, str],
                         output_path: str = None) -> str:
        if not os.path.exists(self.template_path):
            raise FileNotFoundError(f"Template file not found: {self.template_path}")
        
        workbook = load_workbook(self.template_path)
        sheet = workbook.active
        
        for item in data:
            for field, cell_address in column_mappings.items():
                value = item.get(field, '')
                sheet[cell_address] = value
        
        if output_path is None:
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            output_path = f"output_{timestamp}.xlsx"
        
        os.makedirs(os.path.dirname(output_path) if os.path.dirname(output_path) else '.', exist_ok=True)
        workbook.save(output_path)
        
        return output_path
    
    def populate_template_multiple_rows(self, data: List[Dict[str, Any]], 
                                       column_mappings: Dict[str, str],
                                       start_row: int = 2,
                                       output_path: str = None) -> str:
        if not os.path.exists(self.template_path):
            raise FileNotFoundError(f"Template file not found: {self.template_path}")
        
        workbook = load_workbook(self.template_path)
        sheet = workbook.active
        
        for idx, item in enumerate(data):
            current_row = start_row + idx
            for field, col_letter in column_mappings.items():
                if isinstance(col_letter, str) and len(col_letter) <= 3:
                    cell_address = f"{col_letter}{current_row}"
                    value = item.get(field, '')
                    sheet[cell_address] = value
        
        if output_path is None:
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            output_path = f"output_{timestamp}.xlsx"
        
        os.makedirs(os.path.dirname(output_path) if os.path.dirname(output_path) else '.', exist_ok=True)
        workbook.save(output_path)
        
        return output_path
